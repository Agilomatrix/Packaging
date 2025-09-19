import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import hashlib
import tempfile
import shutil
from pathlib import Path
from collections import defaultdict
import zipfile
from PIL import Image
import base64
import traceback
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
import re
from datetime import datetime
from difflib import SequenceMatcher

def navigate_to_step(step_number):
    """Helper function to navigate between steps"""
    if 1 <= step_number <= 6:
        st.session_state.current_step = step_number
        st.rerun()

# Configure Streamlit page
st.set_page_config(
    page_title="AI Packaging Template Mapper",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Initialize session state
if 'current_step' not in st.session_state:
    st.session_state.current_step = 1
if 'selected_packaging_type' not in st.session_state:
    st.session_state.selected_packaging_type = None
if 'template_file' not in st.session_state:
    st.session_state.template_file = None
if 'data_file' not in st.session_state:
    st.session_state.data_file = None
if 'mapping_completed' not in st.session_state:
    st.session_state.mapping_completed = False
if 'image_option' not in st.session_state:
    st.session_state.image_option = None
if 'uploaded_images' not in st.session_state:
    st.session_state.uploaded_images = {}
if 'extracted_excel_images' not in st.session_state:
    st.session_state.extracted_excel_images = {}
if 'all_row_data' not in st.session_state:
    st.session_state.all_row_data = []

def navigate_to_step(step):
    st.session_state.current_step = step
    st.rerun()


class EnhancedImageExtractor:
    """Advanced image extraction and placement with smart positioning"""
    
    def __init__(self):
        self.supported_formats = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        self._placement_counters = defaultdict(int)
        self.current_excel_path = None
        self.row_image_mapping = {}  # Store mapping of rows to images
        
    def analyze_template_structure(self, template_path):
        """Analyze template to find image placement areas"""
        try:
            workbook = openpyxl.load_workbook(template_path, data_only=False)
            worksheet = workbook.active
            
            image_zones = {
                'current_packaging': None,
                'primary_packaging': None,
                'secondary_packaging': None,
                'label': None
            }
            
            # Search for specific headers/keywords in the template
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_value_lower = cell.value.lower()
                        
                        # Look for packaging-related headers
                        if 'current packaging' in cell_value_lower:
                            image_zones['current_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'primary packaging' in cell_value_lower:
                            image_zones['primary_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'secondary packaging' in cell_value_lower:
                            image_zones['secondary_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'label' in cell_value_lower or 'barcode' in cell_value_lower:
                            image_zones['label'] = self._find_image_area_near_cell(worksheet, cell)
            
            workbook.close()
            return image_zones
            
        except Exception as e:
            st.error(f"Error analyzing template structure: {e}")
            return {}
    
    def _find_image_area_near_cell(self, worksheet, header_cell):
        """Find the best area for image placement near a header cell"""
        try:
            header_row = header_cell.row
            header_col = header_cell.column
            
            # Look for merged cells or empty areas below/adjacent to header
            # Strategy 1: Look directly below the header
            for row_offset in range(1, 10):  # Check up to 10 rows below
                target_row = header_row + row_offset
                target_cell = worksheet.cell(target_row, header_col)
                
                # If we find an empty area or a large merged cell, use it
                if not target_cell.value or target_cell.value == "":
                    # Check if this area has enough space (at least 3x3 cells)
                    if self._check_area_availability(worksheet, target_row, header_col, 3, 3):
                        return {
                            'cell': f"{get_column_letter(header_col)}{target_row}",
                            'row': target_row,
                            'col': header_col,
                            'width_cells': 3,
                            'height_cells': 3
                        }
            
            # Strategy 2: Look to the right of header
            for col_offset in range(1, 5):
                target_col = header_col + col_offset
                target_cell = worksheet.cell(header_row, target_col)
                
                if not target_cell.value or target_cell.value == "":
                    if self._check_area_availability(worksheet, header_row, target_col, 3, 3):
                        return {
                            'cell': f"{get_column_letter(target_col)}{header_row}",
                            'row': header_row,
                            'col': target_col,
                            'width_cells': 3,
                            'height_cells': 3
                        }
            
            # Fallback: Use a position relative to header
            return {
                'cell': f"{get_column_letter(header_col)}{header_row + 2}",
                'row': header_row + 2,
                'col': header_col,
                'width_cells': 3,
                'height_cells': 3
            }
            
        except Exception as e:
            st.warning(f"Error finding image area near cell: {e}")
            return None
    
    def _check_area_availability(self, worksheet, start_row, start_col, width, height):
        """Check if an area is available for image placement"""
        try:
            for row in range(start_row, start_row + height):
                for col in range(start_col, start_col + width):
                    cell = worksheet.cell(row, col)
                    if cell.value and str(cell.value).strip():
                        return False
            return True
        except:
            return False
    
    def extract_images_from_excel(self, excel_file_path):
        """Enhanced image extraction with better organization and row mapping"""
        try:
            self.current_excel_path = excel_file_path
            images = {}
            
            st.write("🔍 Extracting images from Excel file...")
            
            # METHOD 1: Standard openpyxl extraction with position detection
            try:
                result1 = self._extract_with_openpyxl_enhanced(excel_file_path)
                images.update(result1)
                st.write(f"✅ Enhanced extraction found {len(result1)} images")
            except Exception as e:
                st.write(f"⚠️ Enhanced extraction failed: {e}")
            
            # METHOD 2: ZIP-based extraction as fallback
            if not images:
                try:
                    result2 = self._extract_with_zipfile_enhanced(excel_file_path)
                    images.update(result2)
                    st.write(f"✅ ZIP extraction found {len(result2)} images")
                except Exception as e:
                    st.write(f"⚠️ ZIP extraction failed: {e}")
            
            if not images:
                st.warning("⚠️ No images found in Excel file.")
            else:
                st.success(f"🎯 Total images extracted: {len(images)}")
                # Build row-to-image mapping
                self._build_row_image_mapping(images)
                # Group images by suspected content
                grouped_images = self._group_images_by_content(images)
                self._display_image_groups(grouped_images)
            
            return {'all_sheets': images}
            
        except Exception as e:
            st.error(f"❌ Error extracting images: {e}")
            return {}

    def _build_row_image_mapping(self, images):
        """Build a mapping of Excel rows to their associated images"""
        self.row_image_mapping = {}
        
        # Group images by row
        for img_key, img_data in images.items():
            row_num = img_data.get('row', 0)
            if row_num > 0:  # Valid row
                if row_num not in self.row_image_mapping:
                    self.row_image_mapping[row_num] = {}
                
                img_type = img_data.get('type', 'current')
                self.row_image_mapping[row_num][img_type] = img_key
        
        st.write(f"📊 Built mapping for {len(self.row_image_mapping)} rows with images")

    def extract_images_for_part(self, data_file, part_number, all_extracted_images, vendor_code, current_row=None):
        """Extract images relevant to a specific part number - FIXED VERSION"""
        try:
            if not all_extracted_images or 'all_sheets' not in all_extracted_images:
                st.warning(f"No images available for part {part_number}")
                return self._get_fallback_images()
            
            all_images = all_extracted_images['all_sheets']
            
            # Strategy 1: Try to use row-specific mapping if available
            if current_row and current_row in self.row_image_mapping:
                row_images = {}
                for img_type, img_key in self.row_image_mapping[current_row].items():
                    if img_key in all_images:
                        row_images[img_key] = all_images[img_key]
                
                if row_images:
                    st.write(f"✅ Found {len(row_images)} row-specific images for part {part_number}")
                    return row_images
            
            # Strategy 2: Use intelligent distribution - give each part a different set
            part_specific_images = self._distribute_images_intelligently(
                all_images, part_number, vendor_code
            )
            
            if not part_specific_images:
                st.warning(f"No specific images found for part {part_number}, using fallback")
                part_specific_images = self._get_fallback_images()
            
            st.write(f"🎯 Assigned {len(part_specific_images)} images to part {part_number}")
            return part_specific_images
            
        except Exception as e:
            st.error(f"Error extracting images for part {part_number}: {e}")
            return self._get_fallback_images()
    
    def _distribute_images_intelligently(self, all_images, part_number, vendor_code):
        """Intelligently distribute images to different parts"""
        try:
            # Group images by type
            images_by_type = {}
            for img_key, img_data in all_images.items():
                img_type = img_data.get('type', 'current')
                if img_type not in images_by_type:
                    images_by_type[img_type] = []
                images_by_type[img_type].append((img_key, img_data))
            
            # Create a hash from part number to ensure consistent but different distribution
            part_hash = hash(part_number) % 1000
            
            selected_images = {}
            
            # For each type, select an image based on the part number
            for img_type, images_list in images_by_type.items():
                if images_list:
                    # Use modulo to cycle through available images
                    index = part_hash % len(images_list)
                    selected_key, selected_data = images_list[index]
                    selected_images[selected_key] = selected_data
                    
                    # Move to next image for next part (simple rotation)
                    part_hash += 1
            
            return selected_images
            
        except Exception as e:
            st.write(f"⚠️ Error in intelligent distribution: {e}")
            return self._get_first_of_each_type(all_images)
    
    def _get_fallback_images(self):
        """Provide fallback when no images are available"""
        st.write("📝 Using placeholder images (no actual images available)")
        return {
            'placeholder_current': {
                'data': None,
                'type': 'current',
                'size': (400, 300),
                'placeholder': True
            },
            'placeholder_primary': {
                'data': None,
                'type': 'primary',
                'size': (400, 300),
                'placeholder': True
            }
        }
    
    def _get_first_of_each_type(self, all_images):
        """Get the first image of each type as fallback"""
        first_images = {}
        seen_types = set()
        
        for img_key, img_data in all_images.items():
            img_type = img_data.get('type', 'current')
            if img_type not in seen_types:
                first_images[img_key] = img_data
                seen_types.add(img_type)
                if len(first_images) >= 4:  # Max 4 types
                    break
        
        return first_images
    
    def _extract_row_number(self, position):
        """Extract row number from cell position like 'A42' -> 42"""
        try:
            import re
            match = re.search(r'(\d+)', position)
            return int(match.group(1)) if match else 0
        except:
            return 0
    
    def _extract_with_openpyxl_enhanced(self, excel_file_path):
        """Enhanced openpyxl extraction with better positioning and row tracking"""
        images = {}
        
        try:
            workbook = openpyxl.load_workbook(excel_file_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                if hasattr(worksheet, '_images') and worksheet._images:
                    for idx, img in enumerate(worksheet._images):
                        try:
                            # Get image data
                            image_data = img._data()
                            image_hash = hashlib.md5(image_data).hexdigest()
                            
                            # Create PIL Image for analysis
                            pil_image = Image.open(io.BytesIO(image_data))
                            
                            # Enhanced position detection
                            position_info = self._get_enhanced_position_info(img, worksheet, idx)
                            
                            # Smart image type classification based on COLUMN, not row
                            image_type = self._smart_classify_image_type_by_column(
                                position_info, worksheet, idx
                            )
                            
                            # Convert to base64
                            buffered = io.BytesIO()
                            pil_image.save(buffered, format="PNG")
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            
                            image_key = f"{image_type}_{sheet_name}_{position_info['position']}_{idx}"
                            images[image_key] = {
                                'data': img_str,
                                'format': 'PNG',
                                'size': pil_image.size,
                                'position': position_info['position'],
                                'row': position_info['row'],  # Important for row tracking
                                'col': position_info['col'],
                                'sheet': sheet_name,
                                'index': idx,
                                'type': image_type,
                                'hash': image_hash,
                                'confidence': position_info.get('confidence', 0.5),
                                'area_context': position_info.get('context', '')
                            }
                            
                        except Exception as e:
                            st.write(f"❌ Failed to extract image {idx} from {sheet_name}: {e}")
            
            workbook.close()
            
        except Exception as e:
            raise Exception(f"Error in enhanced openpyxl extraction: {e}")
        
        return images
    
    def _smart_classify_image_type_by_column(self, position_info, worksheet, idx):
        """Classify image type based on which COLUMN it's in, not row context"""
        try:
            col = position_info['col']
            
            # Check the column header (first few rows of this column)
            for row_num in range(1, 6):  # Check first 5 rows
                try:
                    cell = worksheet.cell(row_num, col + 1)  # +1 because openpyxl is 0-indexed
                    if cell.value and isinstance(cell.value, str):
                        header_text = cell.value.lower().strip()
                        
                        # Match based on column headers from your Excel
                        if 'current packaging' in header_text:
                            return 'current'
                        elif 'primary packaging' in header_text:
                            return 'primary'
                        elif 'secondary packaging' in header_text:
                            return 'secondary'
                        elif 'label' in header_text or 'barcode' in header_text:
                            return 'label'
                except:
                    continue
            
            # Fallback: use column position to guess type
            # Based on typical Excel layout: Current, Primary, Secondary, Label
            if col <= 5:  # Assuming first few columns
                return 'current'
            elif col <= 10:
                return 'primary' 
            elif col <= 15:
                return 'secondary'
            else:
                return 'label'
                
        except Exception as e:
            st.write(f"⚠️ Column classification failed: {e}")
            # Final fallback
            types = ['current', 'primary', 'secondary', 'label']
            return types[idx % len(types)]
    
    def _get_enhanced_position_info(self, img, worksheet, idx):
        """Get enhanced position information including context"""
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from:
                col = anchor._from.col
                row = anchor._from.row
                position = f"{get_column_letter(col + 1)}{row + 1}"
                
                # Analyze surrounding context (but don't rely on it for type classification)
                context = self._analyze_surrounding_context(worksheet, row, col)
                confidence = 0.8 if context else 0.5
                
                return {
                    'position': position,
                    'row': row,
                    'col': col,
                    'context': context,
                    'confidence': confidence
                }
            else:
                return {
                    'position': f"Image_{idx + 1}",
                    'row': 0,
                    'col': 0,
                    'context': '',
                    'confidence': 0.3
                }
                
        except Exception as e:
            return {
                'position': f"Unknown_{idx}",
                'row': 0,
                'col': 0,
                'context': '',
                'confidence': 0.1
            }
    
    def _analyze_surrounding_context(self, worksheet, img_row, img_col):
        """Analyze text around image position to determine its purpose"""
        context_keywords = {
            'current': ['current', 'present', 'existing', 'actual'],
            'primary': ['primary', 'main', 'first', 'initial'],
            'secondary': ['secondary', 'outer', 'external', 'second'],
            'label': ['label', 'barcode', 'sticker', 'tag']
        }
        
        # First, check column headers (more reliable for your structure)
        header_context = self._check_column_headers(worksheet, img_col)
        if header_context:
            return header_context
        
        # Fallback: Check cells in a 5x5 area around the image
        found_context = []
        for row_offset in range(-2, 3):
            for col_offset in range(-2, 3):
                try:
                    cell = worksheet.cell(img_row + row_offset, img_col + col_offset)
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        for context_type, keywords in context_keywords.items():
                            if any(keyword in cell_text for keyword in keywords):
                                found_context.append(context_type)
                except:
                    continue
        
        return ', '.join(set(found_context)) if found_context else ''
    
    def _check_column_headers(self, worksheet, img_col):
        """Check column headers to determine image type based on your Excel structure"""
        try:
            # Check the first few rows for headers
            for row_num in range(1, 5):
                cell = worksheet.cell(row_num, img_col + 1)  # +1 for 1-based indexing
                if cell.value and isinstance(cell.value, str):
                    header_text = cell.value.lower()
                    
                    if 'current packaging' in header_text:
                        return 'current'
                    elif 'primary packaging' in header_text:
                        return 'primary'
                    elif 'secondary packaging' in header_text:
                        return 'secondary'
                    elif 'label' in header_text:
                        return 'label'
            
            return ''
        except:
            return ''
    
    def _extract_with_zipfile_enhanced(self, excel_file_path):
        """Enhanced ZIP extraction with smart organization"""
        images = {}
        
        try:
            with zipfile.ZipFile(excel_file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                
                # Find all image files
                image_files = []
                for f in file_list:
                    if any(f.lower().endswith(ext) for ext in self.supported_formats):
                        image_files.append(f)
                
                # Sort images to maintain consistent order
                image_files.sort()
                
                for idx, image_file in enumerate(image_files):
                    try:
                        with zip_ref.open(image_file) as img_file:
                            image_data = img_file.read()
                            
                            pil_image = Image.open(io.BytesIO(image_data))
                            
                            # Convert to base64
                            buffered = io.BytesIO()
                            pil_image.save(buffered, format="PNG")
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            
                            # Create hash
                            image_hash = hashlib.md5(image_data).hexdigest()
                            
                            # Smart type classification for ZIP-extracted images
                            filename = os.path.basename(image_file).lower()
                            image_type = self._classify_from_filename(filename, idx)
                            
                            image_key = f"{image_type}_ZIP_{filename}_{idx}"
                            
                            images[image_key] = {
                                'data': img_str,
                                'format': 'PNG',
                                'size': pil_image.size,
                                'position': f"ZIP_{idx}",
                                'row': idx + 2,  # Assign sequential rows
                                'col': 0,
                                'sheet': 'ZIP_EXTRACTED',
                                'index': idx,
                                'type': image_type,
                                'hash': image_hash,
                                'source_path': image_file,
                                'confidence': 0.6
                            }
                            
                    except Exception as e:
                        st.write(f"❌ Failed to extract {image_file}: {e}")
        
        except Exception as e:
            raise Exception(f"Error in enhanced ZIP extraction: {e}")
        
        return images
    
    def _classify_from_filename(self, filename, idx):
        """Classify image type based on filename patterns"""
        filename_keywords = {
            'current': ['current', 'present', 'actual', 'now'],
            'primary': ['primary', 'main', 'inner', 'first', '1st'],
            'secondary': ['secondary', 'outer', 'external', 'second', '2nd'],
            'label': ['label', 'barcode', 'tag', 'sticker', 'code']
        }
        
        for img_type, keywords in filename_keywords.items():
            if any(keyword in filename for keyword in keywords):
                return img_type
        
        # Default fallback - cycle through types
        types = ['current', 'primary', 'secondary', 'label']
        return types[idx % len(types)]
    
    def _group_images_by_content(self, images):
        """Group images by their classified types for better organization"""
        grouped = defaultdict(list)
        for key, img_data in images.items():
            img_type = img_data['type']
            grouped[img_type].append((key, img_data))
        return dict(grouped)
    
    def _display_image_groups(self, grouped_images):
        """Display images organized by type"""
        st.subheader("📋 Extracted Images by Type")
        
        for img_type, images_list in grouped_images.items():
            with st.expander(f"{img_type.capitalize()} Images ({len(images_list)} found)"):
                cols = st.columns(min(3, len(images_list)))
                for idx, (key, img_data) in enumerate(images_list):
                    with cols[idx % 3]:
                        st.image(
                            f"data:image/png;base64,{img_data['data']}", 
                            caption=f"{key}\nSize: {img_data['size']}\nConfidence: {img_data.get('confidence', 0.5):.1f}",
                            width=150
                        )
    
    def smart_add_images_to_template(self, template_path, worksheet, uploaded_images):
        """Smart image placement based on template analysis"""
        try:
            added_images = 0
            temp_image_paths = []
            # Analyze template structure
            image_zones = self.analyze_template_structure(template_path)
        
            # Map image types to detected zones
            type_zone_mapping = {
                'current': image_zones.get('current_packaging'),
                'primary': image_zones.get('primary_packaging'),
                'secondary': image_zones.get('secondary_packaging'),
                'label': image_zones.get('label')
            }
        
            for img_key, img_data in uploaded_images.items():
                img_type = img_data.get('type', 'current')
                target_zone = type_zone_mapping.get(img_type)
            
                # Set size based on image type - FIXED SIZE LOGIC
                if img_type == 'current':
                    width_cm, height_cm = 8.3, 8.3  # Current packaging is larger
                else:
                    width_cm, height_cm = 4.3, 4.3  # Primary, secondary, label are smaller
            
                if target_zone:
                    # Use detected zone - but update zone size based on image type
                    if img_type == 'current':
                        # Larger zone for current packaging
                        target_zone['width_cells'] = 4  # Wider for 8.3cm
                        target_zone['height_cells'] = 15  # Taller for 8.3cm
                    else:
                        # Smaller zone for other types
                        target_zone['width_cells'] = 2  # Standard for 4.3cm
                        target_zone['height_cells'] = 7  # Standard for 4.3cm
                    
                    success = self._place_image_smart(
                        worksheet, img_key, img_data, target_zone, temp_image_paths
                    )
                else:
                    # Fallback to default positions with correct sizes
                    fallback_positions = {
                        'current': 'W4',
                        'primary': 'A42',
                        'secondary': 'G42',
                        'label': 'M42'
                    }
                    position = fallback_positions.get(img_type, 'A1')
                    success = self._place_image_at_position(
                        worksheet, img_key, img_data, position,
                        width_cm, height_cm, temp_image_paths  # Use the correct size variables
                    )
            
                if success:
                    added_images += 1
                else:
                    st.write(f"⚠️ Failed to place {img_type} image")
            return added_images, temp_image_paths
        
        except Exception as e:
            st.error(f"Error in smart image placement: {e}")
            return 0, []
    
    def _place_image_smart(self, worksheet, img_key, img_data, zone_info, temp_image_paths):
        """Place image using smart zone information - FIXED VERSION"""
        try:
            # Handle placeholder images
            if img_data.get('placeholder'):
                st.write(f"⏭️ Skipping placeholder image {img_key}")
                return True
             # Create temporary image file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                image_bytes = base64.b64decode(img_data['data'])
                tmp_img.write(image_bytes)
                tmp_img_path = tmp_img.name
            # Create openpyxl image object
            img = OpenpyxlImage(tmp_img_path)
        
            # Calculate size based on zone dimensions
            cell_width_px = 80  # Approximate Excel cell width in pixels
            cell_height_px = 20  # Approximate Excel cell height in pixels
        
            img.width = zone_info['width_cells'] * cell_width_px
            img.height = zone_info['height_cells'] * cell_height_px
        
            # Set position
            img.anchor = zone_info['cell']
        
            # Add image to worksheet
            worksheet.add_image(img)
        
            # Track temporary file for cleanup
            temp_image_paths.append(tmp_img_path)
        
            return True
        except Exception as e:
            st.write(f"❌ Failed to place {img_key} in smart zone: {e}")
            # Clean up temp file if it was created
            if 'tmp_img_path' in locals():
                try:
                    os.unlink(tmp_img_path)
                except:
                    pass
            return False
        
    def add_images_to_template(self, worksheet, uploaded_images):
        """Add uploaded images to template at specific positions - ENHANCED WITH DEBUGGING"""
        try:
            added_images = 0
            temp_image_paths = []
        
            st.write(f"🔍 Starting image placement process with {len(uploaded_images)} images")
        
            # Fixed positions for different image types
            positions = {
                'current': 'W4',      # Current packaging at T3
                'primary': 'A42',     # Primary packaging at A42
                'secondary': 'G42',   # Secondary packaging at F42
                'label': 'M42'        # Label at K42
            }
        
            for img_key, img_data in uploaded_images.items():
                st.write(f"📍 Processing image: {img_key}")
            
                # Skip placeholder images
                if img_data.get('placeholder'):
                    st.write(f"⏭️ Skipping placeholder for {img_key}")
                    continue
                img_type = img_data.get('type', 'current')
                st.write(f"🏷️ Image type: {img_type}")
            
                if img_type in positions:
                    position = positions[img_type]
                    st.write(f"📍 Target position: {position}")
                
                    # Different sizes for different types
                    if img_type == 'current':
                        width_cm, height_cm = 8.3, 8.3
                    else:
                        width_cm, height_cm = 4.3, 4.3
                
                    st.write(f"📏 Image size: {width_cm}x{height_cm} cm")
                
                    success = self._place_image_at_position(
                        worksheet, img_key, img_data, position,
                        width_cm, height_cm, temp_image_paths
                    )
                
                    if success:
                        added_images += 1
                        st.write(f"✅ Successfully added {img_key} to template")
                    else:
                        st.write(f"❌ Failed to add {img_key} to template")
                else:
                    st.write(f"⚠️ Unknown image type: {img_type}")
                st.write(f"📊 Final result: {added_images} images successfully added to template")
                return added_images, temp_image_paths
        except Exception as e:
            st.error(f"❌ Error in add_images_to_template: {e}")
            st.write(f"Error details: {str(e)}")
            return 0, []

    def _place_image_at_position(self, worksheet, img_key, img_data, cell_position, width_cm, height_cm, temp_image_paths):
        """Place a single image at the specified cell position - FIXED VERSION"""
        try:
            # Skip placeholder images
            if img_data.get('placeholder'):
                st.write(f"⏭️ Skipping placeholder image {img_key}")
                return True
            # Create temporary image file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                image_bytes = base64.b64decode(img_data['data'])
                tmp_img.write(image_bytes)
                tmp_img_path = tmp_img.name
            # Create openpyxl image object
            img = OpenpyxlImage(tmp_img_path)
            # Set image size (converting cm to pixels: 1cm ≈ 37.8 pixels)
            img.width = int(width_cm * 37.8)
            img.height = int(height_cm * 37.8)
            # Set position using simple anchor
            img.anchor = cell_position
            # Add image to worksheet
            worksheet.add_image(img)
            # Track temporary file for cleanup
            temp_image_paths.append(tmp_img_path)
            return True
        except Exception as e:
            st.write(f"❌ Failed to place {img_key} at {cell_position}: {e}")
            # Clean up temp file if it was created
            if 'tmp_img_path' in locals():
                try:
                    os.unlink(tmp_img_path)
                except:
                    pass
            return False

            
class EnhancedTemplateMapperWithImages:
    def __init__(self):
        self.image_extractor = EnhancedImageExtractor()
        self.similarity_threshold = 0.3
        
        # Enhanced section-based mapping rules (from your working code)
        self.section_mappings = {
            'general_information': {
                'section_keywords': ['general information', 'document info', 'metadata'],
                'field_mappings': {
                    'date': 'Date',
                    'revision no': 'Revision No.'
                }
            },
            
            'primary_packaging': {
                'section_keywords': [
                    'primary packaging instruction', 'primary packaging', 'primary', 
                    'internal', '( primary / internal )', 'primary / internal'
                ],
                'field_mappings': {
                    'primary packaging type': 'Primary Packaging Type',
                    'packaging type': 'Primary Packaging Type',
                    'l-mm': 'Primary L-mm',
                    'l mm': 'Primary L-mm',
                    'length': 'Primary L-mm',
                    'w-mm': 'Primary W-mm',
                    'w mm': 'Primary W-mm', 
                    'width': 'Primary W-mm',
                    'h-mm': 'Primary H-mm',
                    'h mm': 'Primary H-mm',
                    'height': 'Primary H-mm',
                    'qty/pack': 'Primary Qty/Pack',
                    'quantity': 'Primary Qty/Pack',
                    'empty weight': 'Primary Empty Weight',
                    'pack weight': 'Primary Pack Weight'
                }
            },
            
            'secondary_packaging': {
                'section_keywords': [
                    'secondary packaging instruction', 'secondary packaging', 'secondary', 
                    'outer', 'external', '( outer / external )', 'outer / external'
                ],
                'field_mappings': {
                    'secondary packaging type': 'Secondary Packaging Type',
                    'packaging type': 'Secondary Packaging Type',
                    'type': 'Secondary Packaging Type',
                    'l-mm': 'Secondary L-mm',
                    'l mm': 'Secondary L-mm',
                    'length': 'Secondary L-mm',
                    'w-mm': 'Secondary W-mm',
                    'w mm': 'Secondary W-mm',
                    'width': 'Secondary W-mm',
                    'h-mm': 'Secondary H-mm',
                    'h mm': 'Secondary H-mm',
                    'height': 'Secondary H-mm',
                    'qty/pack': 'Secondary Qty/Pack',
                    'quantity': 'Secondary Qty/Pack',
                    'empty weight': 'Secondary Empty Weight',
                    'pack weight': 'Secondary Pack Weight'
                }
            },
            
            'part_information': {
                'section_keywords': [
                    'part information', 'part info', 'part', 'component', 'item', 'component information'
                ],
                'field_mappings': {
                    'length': 'Part L',
                    'part l': 'Part L',
                    'component l': 'Part L',
                    'W': 'Part W',
                    'w': 'Part W',
                    'width': 'Part W',
                    'part w': 'Part W',
                    'component w': 'Part W',
                    'H': 'Part H',
                    'h': 'Part H',
                    'height': 'Part H',
                    'part h': 'Part H',
                    'component h': 'Part H',
                    'part no': 'Part No',
                    'part number': 'Part No',
                    'description': 'Part Description',
                    'unit weight': 'Part Unit Weight'
                }
            },
            
            'vendor_information': {
                'section_keywords': [
                    'vendor information', 'vendor info', 'vendor', 'supplier', 'supplier information', 'supplier info'
                ],
                'field_mappings': {
                    'vendor name': 'Vendor Name',
                    'name': 'Vendor Name',
                    'supplier name': 'Vendor Name',
                    'vendor code': 'Vendor Code',
                    'supplier code': 'Vendor Code',
                    'code': 'Vendor Code',
                    'vendor location': 'Vendor Location',
                    'location': 'Vendor Location',
                    'supplier location': 'Vendor Location',
                    'address': 'Vendor Location'
                }
            },
            
            'procedure_information': {
                'section_keywords': [
                    'procedure information', 'procedure', 'packaging procedure', 'loading details',
                    'pallet information', 'pallet details', 'packaging details',
                    'loading instruction', 'packing procedure', 'palletization'
                ],
                'field_mappings': {
                    # x No. of Parts mapping (Column AH: "x No. of Parts")
                    'x no. of parts': 'x No. of Parts',
                    'x no of parts': 'x No. of Parts',
                    'x number of parts': 'x No. of Parts',
                    'no. of parts': 'x No. of Parts',
                    'no of parts': 'x No. of Parts',
                    'number of parts': 'x No. of Parts',
                    'parts': 'x No. of Parts',
                
                    # Layer mapping (Column AF: "Layer") 
                    'layer': 'Layer',
                    'layers': 'Layer',
                    'max layer': 'Layer',
                    'maximum layer': 'Layer',
                    'pallet layer': 'Layer',
                    'boxes per layer': 'Layer',
                
                    # Level mapping (Column AG: "Level")
                    'level': 'Level',
                    'levels': 'Level',
                    'max level': 'Level',
                    'maximum level': 'Level',
                    'stacking level': 'Level',
                    'pallet level': 'Level',
                
                    # Inner/Outer dimensions
                    'inner l': 'Inner L',
                    'inner w': 'Inner W', 
                    'inner h': 'Inner H',
                    'inner length': 'Inner L',
                    'inner width': 'Inner W',
                    'inner height': 'Inner H',
                    'outer l': 'Outer L',
                    'outer w': 'Outer W',
                    'outer h': 'Outer H',
                    'outer length': 'Outer L',
                    'outer width': 'Outer W',
                    'outer height': 'Outer H',
                    'inner qty/pack': 'Inner Qty/Pack',
                    'inner quantity': 'Inner Qty/Pack',
                    'inner qty': 'Inner Qty/Pack'
                }
            },
            'miscellaneous_information': {
                'section_keywords': ['problems if any', 'remarks', 'notes'],
                'field_mappings': {
                    'problems if any': 'Problems',  # Maps template label to data column "Problems"
                    'problems': 'Problems',
                    'remarks': 'Remarks'
                }
            }
        }

    def preprocess_text(self, text):
        """Preprocess text for better matching"""
        try:
            if pd.isna(text) or text is None:
                return ""
            
            text = str(text).lower()
            text = re.sub(r'[()[\]{}]', ' ', text)
            text = re.sub(r'[^\w\s/-]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            
            return text
        except Exception as e:
            st.error(f"Error in preprocess_text: {e}")
            return ""

    def is_mappable_field(self, text):
        """Enhanced field detection for packaging templates"""
        try:
            if not text or pd.isna(text):
                return False
            text = str(text).lower().strip()
            if not text:
                return False
            print(f"DEBUG is_mappable_field: Checking '{text}'")

            # Skip header-like patterns that should not be treated as fields
            header_exclusions = [
                'vendor information', 'part information', 'primary packaging', 'secondary packaging',
                'packaging instruction', 'procedure', 'steps', 'process'
            ]
            for exclusion in header_exclusions:
                if exclusion in text and 'type' not in text:
                    print(f"DEBUG: Excluding '{text}' as header")
                    return False
        
            # Define mappable field patterns for packaging templates
            mappable_patterns = [
                r'primary\s+packaging\s+type', r'secondary\s+packaging\s+type', 
                r'packaging\s+type', r'\btype\b',
                r'\bl[-\s]*mm\b', r'\bw[-\s]*mm\b', r'\bh[-\s]*mm\b',
                r'\bl\b', r'\bw\b', r'\bh\b',
                r'part\s+l\b', r'part\s+w\b', r'part\s+h\b',
                r'\blength\b', r'\bwidth\b', r'\bheight\b',
                r'qty[/\s]*pack', r'quantity\b', r'weight\b', r'empty\s+weight',
                r'\bcode\b', r'\bname\b', r'\bdescription\b', r'\blocation\b',
                r'part\s+no\b', r'part\s+number\b',
                r'\bdate\b',
                r'\brev(ision)?\s*no\.?\b',
                # Procedure-specific patterns
                r'\bx\s*no\.?\s*of\s*parts\b',
                r'\bx\s*no\s*of\s*parts\b',
                r'\bx\s*number\s*of\s*parts\b',
                r'\bno\.?\s*of\s*parts\b',
                r'\bnumber\s*of\s*parts\b',
                r'\bparts\s*per\s*pack\b',
                r'\bparts\s*quantity\b',
                r'\bqty\s*of\s*parts\b',
                r'\blevel\b', r'\blevels\b',
                r'\blayer\b', r'\blayers\b',
                r'\bmax\s*level\b', r'\bmaximum\s*level\b',
                r'\bmax\s*layer\b', r'\bmaximum\s*layer\b',
                r'\bstacking\s*level\b', r'\bpallet\s*level\b',
                r'\binner\s*l\b', r'\binner\s*length\b',
                r'\binner\s*w\b', r'\binner\s*width\b', 
                r'\binner\s*h\b', r'\binner\s*height\b',
                r'\binner\s*qty[/\s]*pack\b',
                r'\bouter\s*l\b', r'\bouter\s*length\b',
                r'\bouter\s*w\b', r'\bouter\s*width\b',
                r'\bouter\s*h\b', r'\bouter\s*height\b',
                r'\bpallet\b', r'\bpalletiz\w*\b',
                r'\bproblems\b' 
            ]
        
            for pattern in mappable_patterns:
                if re.search(pattern, text):
                    print(f"DEBUG: '{text}' matches pattern '{pattern}'")
                    return True
        
            if text.endswith(':'):
                print(f"DEBUG: '{text}' ends with colon")
                return True
        
            print(f"DEBUG: '{text}' is NOT mappable")
            return False
        except Exception as e:
            st.error(f"Error in is_mappable_field: {e}")
            return False

    def identify_section_context(self, worksheet, row, col, max_search_rows=15):
        """Enhanced section identification with better pattern matching"""
        try:
            section_context = None
            for search_row in range(max(1, row - max_search_rows), row + 5):
                for search_col in range(max(1, col - 20), min(worksheet.max_column + 1, col + 20)):
                    try:
                        cell = worksheet.cell(row=search_row, column=search_col)
                        if cell.value:
                            cell_text = self.preprocess_text(str(cell.value))
                    
                            for section_name, section_info in self.section_mappings.items():
                                for keyword in section_info['section_keywords']:
                                    keyword_processed = self.preprocess_text(keyword)
                            
                                    if keyword_processed == cell_text or keyword_processed in cell_text or cell_text in keyword_processed:
                                        print(f"DEBUG: Found section context '{section_name}' for field at ({row}, {col}) via keyword '{keyword}'")
                                        return section_name
                            
                                # Enhanced context matching
                                if section_name == 'procedure_information':
                                    procedure_indicators = [
                                        'procedure', 'loading', 'pallet', 'packaging procedure',
                                        'stacking', 'palletization', 'loading details', 
                                        'packing instruction', 'step', 'layer', 'level'
                                    ]
                                    if any(indicator in cell_text for indicator in procedure_indicators):
                                        print(f"DEBUG: Found procedure context for field at ({row}, {col}) via indicator in '{cell_text}'")
                                        return section_name
                            
                                elif section_name == 'primary_packaging':
                                    if ('primary' in cell_text and ('packaging' in cell_text or 'internal' in cell_text)):
                                        return section_name
                                elif section_name == 'secondary_packaging':
                                    if ('secondary' in cell_text and ('packaging' in cell_text or 'outer' in cell_text or 'external' in cell_text)):
                                        return section_name
                                elif section_name == 'part_information':
                                    if (('part' in cell_text and ('information' in cell_text or 'info' in cell_text)) or ('component' in cell_text and ('information' in cell_text or 'info' in cell_text))):
                                        return section_name
                                elif section_name == 'vendor_information':
                                    if (('vendor' in cell_text and ('information' in cell_text or 'info' in cell_text)) or ('supplier' in cell_text and ('information' in cell_text or 'info' in cell_text))):
                                        return section_name
                    except:
                        continue
    
            return self.infer_section_from_field_name(row, col)
    
        except Exception as e:
            st.error(f"Error in identify_section_context: {e}")
            return None
        
    def infer_section_from_field_name(self, row, col):
        """Infer section context from field name when no explicit section header found"""
        try:
            return 'procedure_information'
        except Exception as e:
            return 'procedure_information'

    def calculate_similarity(self, text1, text2):
        """Calculate similarity between two texts"""
        try:
            if not text1 or not text2:
                return 0.0
            
            text1 = self.preprocess_text(text1)
            text2 = self.preprocess_text(text2)
            
            if not text1 or not text2:
                return 0.0
            
            sequence_sim = SequenceMatcher(None, text1, text2).ratio()
            return sequence_sim
        except Exception as e:
            st.error(f"Error in calculate_similarity: {e}")
            return 0.0

    def find_template_fields_with_context_and_images(self, template_file):
        """Find template fields and image upload areas"""
        fields = {}
        image_areas = []
        try:
            workbook = openpyxl.load_workbook(template_file)
            worksheet = workbook.active
        
            merged_ranges = worksheet.merged_cells.ranges
        
            for row in worksheet.iter_rows():
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_value = str(cell.value).strip()

                            # Force capture Date & Revision No anywhere in sheet
                            if cell_value.lower() in ['date', 'revision no.', 'revision no']:
                                fields[cell.coordinate] = {
                                    'value': cell_value,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'merged_range': None,
                                    'section_context': 'general_information',
                                    'is_mappable': True
                                }
                                continue
                        
                            if cell_value and self.is_mappable_field(cell_value):
                                cell_coord = cell.coordinate
                                merged_range = None
                            
                                for merge_range in merged_ranges:
                                    if cell.coordinate in merge_range:
                                        merged_range = str(merge_range)
                                        break
                            
                                section_context = self.identify_section_context(
                                    worksheet, cell.row, cell.column
                                )
                            
                                print(f"DEBUG: Found field '{cell_value}' at {cell_coord}")
                                print(f"DEBUG: Section context: {section_context}")
                                print("---")
                            
                                fields[cell_coord] = {
                                    'value': cell_value,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'merged_range': merged_range,
                                    'section_context': section_context,
                                    'is_mappable': True
                                }
                    except Exception as e:
                        continue
        
            workbook.close()
        
        except Exception as e:
            st.error(f"Error reading template: {e}")
    
        return fields, image_areas

    def find_data_cell_for_label(self, worksheet, field_info):
        """
        Find data cell for a label, with special handling for the 'Problems' field
        and safer, constrained search for all other fields.
        """
        try:
            row = field_info['row']
            col = field_info['column']
            field_text_lower = self.preprocess_text(field_info.get('value', ''))

            # --- SPECIAL CASE: Handle "PROBLEMS IF ANY" field directly ---
            if 'problems' in field_text_lower:
                # Based on the template, the label is at V23, and the green cell starts at V25.
                # This is a fixed position relative to the label: (row + 2, same column).
                target_cell_coord = worksheet.cell(row=row + 2, column=col).coordinate
                st.write(f"INFO: Found 'Problems' field. Targeting specific cell: {target_cell_coord}")
                return target_cell_coord

            # --- STANDARD PLACEMENT LOGIC FOR ALL OTHER FIELDS ---

            # Define columns to be explicitly ignored for general placement.
            # The special case above runs first, so this won't block the 'Problems' field.
            IGNORED_COLUMNS = [22, 23, 24, 25] # V, W, X, Y

            def is_suitable_data_cell(r, c):
                """Check if a cell at a given row and column is suitable for data entry."""
                if not (1 <= r <= worksheet.max_row and 1 <= c <= worksheet.max_column):
                    return False
                if c in IGNORED_COLUMNS:
                    return False
                try:
                    cell = worksheet.cell(row=r, column=c)
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        return False
                    if cell.value is None or str(cell.value).strip() == "":
                        return True
                    cell_text = str(cell.value).lower().strip()
                    data_patterns = [r'^_+$', r'^\.*$', r'^-+$', r'enter', r'fill', r'data']
                    return any(re.search(pattern, cell_text) for pattern in data_patterns)
                except:
                    return False

            # Strategy 1: Look RIGHT of the label (up to 5 cells)
            for offset in range(1, 6):
                if is_suitable_data_cell(row, col + offset):
                    return worksheet.cell(row=row, column=col + offset).coordinate
            
            # Strategy 2: Look IMMEDIATELY BELOW the label (1 cell)
            if is_suitable_data_cell(row + 1, col):
                return worksheet.cell(row=row + 1, column=col).coordinate

            # If no suitable cell is found, give up to prevent errors.
            st.write(f"WARNING: Could not find a safe data cell for label '{field_info['value']}'. Skipping placement.")
            return None
            
        except Exception as e:
            st.error(f"Error in find_data_cell_for_label for '{field_info.get('value', 'N/A')}': {e}")
            return None

    # *** NEW METHOD: Read procedure steps from Excel template ***
    def read_procedure_steps_from_template(self, template_path, packaging_type=None):
        """
        Final robust version: Reads steps only from the clearly defined 'Packaging Procedure' block.
        It finds the start and end of the block and ignores everything outside of it.
        This prevents duplicate steps and the inclusion of unwanted labels.
        """
        try:
            print("\n=== READING PROCEDURE STEPS (BLOCK-AWARE VERSION) ===")
            st.write(f"📖 Reading procedure steps from template...")
    
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
    
            procedure_steps = []
            start_row = -1
            end_row = -1
            
            # --- STEP 1: Find the start and end row numbers of the procedure block ---
            for row in worksheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=5):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text_lower = cell.value.lower().strip()
                        if "packaging procedure" in cell_text_lower:
                            start_row = cell.row
                        # Use a very specific stop phrase from your screenshot
                        elif "reference image/pictures" in cell_text_lower:
                            end_row = cell.row
                            break
                if end_row != -1:
                    break
            
            if start_row == -1:
                st.warning("Could not find the 'Packaging Procedure' start header in the template.")
                return []
            if end_row == -1:
                st.warning("Could not find the 'Reference Image/Pictures' end header. This might lead to extra steps.")
                end_row = start_row + 25 # Fallback to prevent an infinite loop

            print(f"✅ Procedure block identified: Reading from row {start_row + 1} up to row {end_row - 1}")

            # --- STEP 2: Iterate ONLY within the identified block ---
            # For each row, we ONLY check the cell in column B (index 2). This prevents duplicates.
            for row_num in range(start_row + 1, end_row):
                # Target the specific cell where the step text begins
                cell = worksheet.cell(row=row_num, column=2)
                
                if cell.value and isinstance(cell.value, str):
                    step_text = cell.value.strip()
                    
                    # --- FINAL FILTER ---
                    # 1. Ensure text is not empty.
                    # 2. Ensure text has more than 3 words (this filters out labels like "Secondary Packaging").
                    if step_text and len(step_text.split()) > 3:
                        procedure_steps.append(step_text)
                        print(f"📝 Found step {len(procedure_steps)}: {step_text[:60]}...")

            workbook.close()
    
            print(f"✅ FINAL COUNT: Successfully read {len(procedure_steps)} procedure steps.")
            st.write(f"✅ Found {len(procedure_steps)} procedure steps in template")
    
            return procedure_steps
        except Exception as e:
            print(f"❌ Error reading procedure steps from template: {e}")
            st.error(f"Error reading procedure steps from template: {e}")
            return []
            
    def substitute_placeholders_in_steps(self, procedure_steps, data_dict):
        """
        Replace placeholders in procedure steps with actual data values.
        
        Args:
            procedure_steps: List of steps with {placeholders}
            data_dict: Dictionary containing mapped data values
            
        Returns:
            List of procedure steps with placeholders replaced
        """
        try:
            print(f"\n=== SUBSTITUTING PLACEHOLDERS IN STEPS ===")
            st.write(f"🔄 Replacing placeholders with actual data...")
            
            # Debug: Print available data
            print(f"Available data in data_dict:")
            for key, value in data_dict.items():
                print(f"  '{key}': '{value}'")
            print("=" * 50)
            
            filled_steps = []
            
            for i, step in enumerate(procedure_steps, 1):
                filled_step = step
                
                print(f"Processing step {i}: {step[:50]}...")
                
                # Enhanced mapping with multiple fallback options
                replacements = {
                    # *** CRITICAL: Enhanced quantity mappings - multiple fallbacks ***
                    '{x No. of Parts}': (
                        data_dict.get('x No. of Parts') or 
                        data_dict.get('X No. of Parts') or
                        data_dict.get('x no. of parts') or
                        data_dict.get('X no. of parts') or
                        data_dict.get('no. of parts') or
                        data_dict.get('No. of Parts') or
                        data_dict.get('number of parts') or
                        data_dict.get('Number of Parts') or
                        data_dict.get('parts per pack') or
                        data_dict.get('Parts Per Pack') or
                        data_dict.get('qty of parts') or
                        data_dict.get('Qty of Parts') or
                        '8'  # Default fallback
                    ),
                
                    # *** CRITICAL: Enhanced Level mappings - multiple fallbacks ***
                    '{Level}': (
                        data_dict.get('Level') or
                        data_dict.get('level') or
                        data_dict.get('LEVEL') or
                        data_dict.get('Levels') or
                        data_dict.get('levels') or
                        data_dict.get('max level') or
                        data_dict.get('Max Level') or
                        data_dict.get('maximum level') or
                        data_dict.get('Maximum Level') or
                        data_dict.get('stacking level') or
                        data_dict.get('Stacking Level') or
                        '5'  # Default fallback
                    ),
                
                    # *** CRITICAL: Enhanced Layer mappings - multiple fallbacks ***
                    '{Layer}': (
                        data_dict.get('Layer') or
                        data_dict.get('layer') or
                        data_dict.get('LAYER') or
                        data_dict.get('Layers') or
                        data_dict.get('layers') or
                        data_dict.get('max layer') or
                        data_dict.get('Max Layer') or
                        data_dict.get('maximum layer') or
                        data_dict.get('Maximum Layer') or
                        '4'  # Default fallback
                    ),
                    
                    # ========= FIX STARTS HERE =========
                    # Added robust placeholder substitution for Primary and Secondary Packaging Types
                    '{Primary Packaging Type}': (
                        data_dict.get('Primary Packaging Type') or
                        data_dict.get('primary packaging type') or
                        data_dict.get('Packaging Type') or  # Fallback to a generic name
                        'N/A'
                    ),
                    '{Secondary Packaging Type}': (
                        data_dict.get('Secondary Packaging Type') or
                        data_dict.get('secondary packaging type') or
                        'N/A'
                    ),
                    # ========= FIX ENDS HERE =========
                    
                    # Inner dimensions - try multiple key variations
                    '{Inner L}': (
                        data_dict.get('Inner L') or 
                        data_dict.get('inner l') or
                        data_dict.get('Inner l') or
                        data_dict.get('INNER L') or
                        data_dict.get('Inner Length') or
                        data_dict.get('inner length') or
                        'XXX'
                    ),
                    '{Inner W}': (
                        data_dict.get('Inner W') or 
                        data_dict.get('inner w') or
                        data_dict.get('Inner w') or
                        data_dict.get('INNER W') or
                        data_dict.get('Inner Width') or
                        data_dict.get('inner width') or
                        'XXX'
                    ),
                    '{Inner H}': (
                        data_dict.get('Inner H') or 
                        data_dict.get('inner h') or
                        data_dict.get('Inner h') or
                        data_dict.get('INNER H') or
                        data_dict.get('Inner Height') or
                        data_dict.get('inner height') or
                        'XXX'
                    ),
                    
                    # Inner Qty/Pack - try multiple variations
                    '{Inner Qty/Pack}': (
                        data_dict.get('Inner Qty/Pack') or
                        data_dict.get('inner qty/pack') or
                        data_dict.get('Inner qty/pack') or
                        data_dict.get('INNER QTY/PACK') or
                        data_dict.get('Inner Quantity') or
                        data_dict.get('inner quantity') or
                        '1'
                    ),
                    
                    # Outer dimensions - try multiple variations
                    '{Outer L}': (
                        data_dict.get('Outer L') or 
                        data_dict.get('outer l') or
                        data_dict.get('Outer l') or
                        data_dict.get('OUTER L') or
                        data_dict.get('Outer Length') or
                        data_dict.get('outer length') or
                        'XXX'
                    ),
                    '{Outer W}': (
                        data_dict.get('Outer W') or 
                        data_dict.get('outer w') or
                        data_dict.get('Outer w') or
                        data_dict.get('OUTER W') or
                        data_dict.get('Outer Width') or
                        data_dict.get('outer width') or
                        'XXX'
                    ),
                    '{Outer H}': (
                        data_dict.get('Outer H') or 
                        data_dict.get('outer h') or
                        data_dict.get('Outer h') or
                        data_dict.get('OUTER H') or
                        data_dict.get('Outer Height') or
                        data_dict.get('outer height') or
                        'XXX'
                    ),
                    
                    # Primary Qty/Pack - try multiple variations
                    '{Primary Qty/Pack}': (
                        data_dict.get('Primary Qty/Pack') or
                        data_dict.get('primary qty/pack') or
                        data_dict.get('Primary qty/pack') or
                        data_dict.get('PRIMARY QTY/PACK') or
                        data_dict.get('Primary Quantity') or
                        data_dict.get('primary quantity') or
                        '1'
                    ),
                    
                    # Generic Qty/Pack - try multiple variations
                    '{Qty/Pack}': (
                        data_dict.get('Qty/Pack') or
                        data_dict.get('qty/pack') or
                        data_dict.get('QTY/PACK') or
                        data_dict.get('Quantity') or
                        data_dict.get('quantity') or
                        '1'
                    ),
                    '{Qty/Veh}': (
                        data_dict.get('Qty/Veh') or
                        data_dict.get('qty/veh') or
                        data_dict.get('QTY/VEH') or
                        data_dict.get('Qty/Pack') or
                        data_dict.get('qty/pack') or
                        '1'
                    ),
                    
                    # Secondary dimensions
                    '{Secondary L-mm}': (
                        data_dict.get('Secondary L-mm') or
                        data_dict.get('secondary l-mm') or
                        data_dict.get('Secondary L') or
                        data_dict.get('secondary l') or
                        'XXX'
                    ),
                    '{Secondary W-mm}': (
                        data_dict.get('Secondary W-mm') or
                        data_dict.get('secondary w-mm') or
                        data_dict.get('Secondary W') or
                        data_dict.get('secondary w') or
                        'XXX'
                    ),
                    '{Secondary H-mm}': (
                        data_dict.get('Secondary H-mm') or
                        data_dict.get('secondary h-mm') or
                        data_dict.get('Secondary H') or
                        data_dict.get('secondary h') or
                        'XXX'
                    ),
                    
                    # Primary dimensions
                    '{Primary L-mm}': (
                        data_dict.get('Primary L-mm') or
                        data_dict.get('primary l-mm') or
                        data_dict.get('Primary L') or
                        data_dict.get('primary l') or
                        'XXX'
                    ),
                    '{Primary W-mm}': (
                        data_dict.get('Primary W-mm') or
                        data_dict.get('primary w-mm') or
                        data_dict.get('Primary W') or
                        data_dict.get('primary w') or
                        'XXX'
                    ),
                    '{Primary H-mm}': (
                        data_dict.get('Primary H-mm') or
                        data_dict.get('primary h-mm') or
                        data_dict.get('Primary H') or
                        data_dict.get('primary h') or
                        'XXX'
                    )
                }
                
                # Debug: Show what replacements are being made
                for placeholder, raw_value in replacements.items():
                    if placeholder in filled_step:
                        clean_value = self.clean_data_value(raw_value)
                        if not clean_value or clean_value == "":
                            clean_value = 'XXX'
                        print(f"  Replacing {placeholder} with '{clean_value}' (from: {raw_value})")
                        filled_step = filled_step.replace(placeholder, str(clean_value))
                
                filled_steps.append(filled_step)
                print(f"  Final step {i}: {filled_step[:100]}...")
                print("---")
            
            print(f"✅ Successfully processed {len(filled_steps)} procedure steps")
            st.write(f"✅ Replaced placeholders in {len(filled_steps)} steps")
            
            return filled_steps
            
        except Exception as e:
            print(f"❌ Error substituting placeholders: {e}")
            st.error(f"Error substituting placeholders: {e}")
            return procedure_steps  # Return original steps if substitution fails

    def map_data_with_section_context(self, template_fields, data_df):
        """Enhanced mapping with EXACT column name matching"""
        mapping_results = {}
        used_columns = set()

        try:
            data_columns = data_df.columns.tolist()
            print(f"DEBUG: Available data columns: {data_columns}")
        
            for coord, field in template_fields.items():
                try:
                    best_match = None
                    best_score = 0.0
                    field_value = field['value']
                    section_context = field.get('section_context')

                    print(f"DEBUG: Mapping field '{field_value}' with section '{section_context}'")

                    # Direct exact column name matching first
                    field_lower = self.preprocess_text(field_value)
                
                    for data_col in data_columns:
                        if data_col in used_columns:
                            continue
                        
                        col_lower = self.preprocess_text(data_col)
                    
                        # Exact matches
                        if col_lower == field_lower:
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT EXACT MATCH: '{field_value}' → '{data_col}'")
                            break
                    
                        # Special case matches for specific columns
                        if field_lower == 'layer' and col_lower == 'layer':
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT LAYER MATCH: '{field_value}' → '{data_col}'")
                            break
                        elif field_lower == 'level' and col_lower == 'level':
                            best_match = data_col  
                            best_score = 1.0
                            print(f"DEBUG: DIRECT LEVEL MATCH: '{field_value}' → '{data_col}'")
                            break
                        elif ('x no' in field_lower or 'no. of parts' in field_lower) and 'x no of parts' in col_lower:
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT X NO OF PARTS MATCH: '{field_value}' → '{data_col}'")
                            break

                    # If direct match found, use it
                    if best_match:
                        mapping_results[coord] = {
                            'template_field': field_value,
                            'data_column': best_match,
                            'similarity': best_score,
                            'field_info': field,
                            'section_context': section_context,
                            'is_mappable': True
                        }
                        used_columns.add(best_match)
                        print(f"DEBUG: DIRECT MATCH SUCCESS: {field_value} → {best_match}")
                        continue

                    # Force procedure context for specific fields
                    if not section_context:
                        procedure_fields = ['layer', 'level', 'x no of parts', 'no. of parts', 'parts']
                        if any(proc_field in field_lower for proc_field in procedure_fields):
                            section_context = 'procedure_information'
                            print(f"DEBUG: FORCED procedure context for field '{field_value}'")

                    # Section mapping logic
                    if section_context and section_context in self.section_mappings:
                        section_mappings = self.section_mappings[section_context]['field_mappings']
                        print(f"DEBUG: Section mappings: {section_mappings}")

                        for template_field_key, data_column_pattern in section_mappings.items():
                            normalized_field_value = self.preprocess_text(field_value)
                            normalized_template_key = self.preprocess_text(template_field_key)

                            print(f"DEBUG: Comparing '{normalized_field_value}' with '{normalized_template_key}'")

                            if normalized_field_value == normalized_template_key:
                                # For procedure_information, don't add section prefix
                                if section_context == "procedure_information":
                                    expected_column = data_column_pattern 
                                else:
                                    section_prefix = section_context.split('_')[0].capitalize()
                                    expected_column = f"{section_prefix} {data_column_pattern}".strip()
                            
                                print(f"DEBUG: Looking for expected column: '{expected_column}'")

                                for data_col in data_columns:
                                    if data_col in used_columns:
                                        continue
                                    if self.preprocess_text(data_col) == self.preprocess_text(expected_column):
                                        best_match = data_col
                                        best_score = 1.0
                                        print(f"DEBUG: SECTION EXACT MATCH FOUND: {data_col}")
                                        break

                                # Fallback to similarity match if no exact match
                                if not best_match:
                                    for data_col in data_columns:
                                        if data_col in used_columns:
                                            continue
                                        similarity = self.calculate_similarity(expected_column, data_col)
                                        if similarity > best_score and similarity >= self.similarity_threshold:
                                            best_score = similarity
                                            best_match = data_col
                                            print(f"DEBUG: SECTION SIMILARITY MATCH: {data_col} (score: {similarity})")
                                break

                    # Final fallback: general similarity matching
                    if not best_match:
                        for data_col in data_columns:
                            if data_col in used_columns:
                                continue
                            similarity = self.calculate_similarity(field_value, data_col)
                            if similarity > best_score and similarity >= self.similarity_threshold:
                                best_score = similarity
                                best_match = data_col

                    print(f"DEBUG: Final mapping result - Field: '{field_value}' → Column: '{best_match}' (Score: {best_score})")
                    print("=" * 50)

                    # Save mapping
                    mapping_results[coord] = {
                        'template_field': field_value,
                        'data_column': best_match,
                        'similarity': best_score,
                        'field_info': field,
                        'section_context': section_context,
                        'is_mappable': best_match is not None
                    }

                    if best_match:
                        used_columns.add(best_match)

                except Exception as e:
                    st.error(f"Error mapping field {coord}: {e}")
                    continue

        except Exception as e:
            st.error(f"Error in map_data_with_section_context: {e}")
        return mapping_results

    def clean_data_value(self, value):
        """Clean data value to handle NaN, None, and empty values"""
        if pd.isna(value) or value is None:
            return ""
        
        str_value = str(value).strip()
        
        if str_value.lower() in ['nan', 'none', 'null', 'n/a', '#n/a', '']:
            return ""
            
        return str_value

    def map_template_with_data(self, template_path, data_path):
        """Enhanced mapping with section-based approach and multiple row processing"""
        try:
            # Read data from Excel with proper NaN handling
            data_df = pd.read_excel(data_path)
            data_df = data_df.fillna("")
            st.write(f"📊 Loaded data with {len(data_df)} rows and {len(data_df.columns)} columns")
            
            # *** NEW: Read procedure steps from template ONCE ***
            template_procedure_steps = self.read_procedure_steps_from_template(template_path)
            if not template_procedure_steps:
                st.warning("⚠️ No procedure steps found in template. Will use empty steps.")
            
            # Store all row data for multi-template generation
            st.session_state.all_row_data = []
    
            # Process each row
            for row_idx in range(len(data_df)):
                st.write(f"🔄 Processing row {row_idx + 1}/{len(data_df)}")
                
                # Load fresh template for each row
                workbook = openpyxl.load_workbook(template_path)
                worksheet = workbook.active
        
                # Find template fields with section context
                template_fields, _ = self.find_template_fields_with_context_and_images(template_path)
        
                # Map data with section context for current row
                mapping_results = self.map_data_with_section_context_for_row(template_fields, data_df, row_idx)
        
                # Apply mappings to template
                mapping_count = 0
                data_dict = {}  # Store mapped data for procedure generation
                filename_parts = {}  # Store parts for filename
                
                # =================== FIX STARTS HERE ===================
                # Pre-load critical data for procedure steps, regardless of whether a matching field
                # exists in the template. This ensures data is always available for substitution.
                st.write("...Pre-loading critical data for procedures...")
                critical_data_map = {
                    'Secondary Packaging Type': ['Secondary Packaging Type', 'secondary packaging type'],
                    'Primary Packaging Type': ['Primary Packaging Type', 'primary packaging type'],
                    'Outer L': ['Outer L', 'outer l', 'Outer Length', 'outer length', 'Outer L-mm'],
                    'Outer W': ['Outer W', 'outer w', 'Outer Width', 'outer width', 'Outer W-mm'],
                    'Outer H': ['Outer H', 'outer h', 'Outer Height', 'outer height', 'Outer H-mm'],
                }

                for canonical_name, possible_names in critical_data_map.items():
                    for data_col in data_df.columns:
                        # Find the first matching column in the data file
                        if self.preprocess_text(data_col) in [self.preprocess_text(p) for p in possible_names]:
                            raw_value = data_df[data_col].iloc[row_idx]
                            data_dict[canonical_name] = self.clean_data_value(raw_value)
                            print(f"DEBUG: Pre-loaded '{canonical_name}' from column '{data_col}' with value '{data_dict[canonical_name]}'")
                            break  # Move to the next canonical name once found
                # =================== FIX ENDS HERE =====================
        
                for coord, mapping in mapping_results.items():
                    if mapping['is_mappable'] and mapping['data_column']:
                        try:
                            data_col = mapping['data_column']
                            raw_value = data_df[data_col].iloc[row_idx]  # Use current row
                            data_value = self.clean_data_value(raw_value)
                    
                            # Store in data_dict for procedure generation
                            template_field_key = mapping.get('template_field', '').strip()
                            data_dict[template_field_key] = data_value

                            # Also store under the 'canonical' name from section_mappings for robustness
                            section = mapping.get('section_context')
                            if section and section in self.section_mappings:
                                for map_key, canonical_name in self.section_mappings[section]['field_mappings'].items():
                                    if self.preprocess_text(template_field_key) == self.preprocess_text(map_key):
                                        data_dict[canonical_name] = data_value
                                        break
                    
                            # Store filename components
                            data_col_name = mapping.get('data_column', '').lower()
                            if data_col_name:
                                if 'part_no' not in filename_parts and any(term in data_col_name for term in ['part no', 'part_no', 'part number', 'part_number', 'part #']):
                                    filename_parts['part_no'] = data_value
                                if 'description' not in filename_parts and any(term in data_col_name for term in ['description', 'desc', 'part desc']):
                                    filename_parts['description'] = data_value
                                if 'vendor_code' not in filename_parts and any(term in data_col_name for term in ['vendor code', 'vendor_code', 'supplier code']):
                                    filename_parts['vendor_code'] = data_value
                    
                            # Find target cell and write data
                            target_cell_coord = self.find_data_cell_for_label(worksheet, mapping['field_info'])
                    
                            if target_cell_coord and data_value:
                                target_cell = worksheet[target_cell_coord]
                                target_cell.value = data_value
                                mapping_count += 1
                        except Exception as e:
                            st.write(f"⚠️ Error processing row {row_idx + 1}, field '{mapping['template_field']}': {e}")
                
                # *** Process procedure steps from template ***
                steps_written = 0
                if template_procedure_steps:
                    # Substitute placeholders with actual data
                    filled_steps = self.substitute_placeholders_in_steps(template_procedure_steps, data_dict)
                    
                    # Write the filled steps back to template
                    steps_written = self.write_filled_steps_to_template(worksheet, filled_steps)
                else:
                    st.write("⚠️ No procedure steps to process for this row")
                
                # Generate filename
                vendor_code = filename_parts.get('vendor_code', 'NoVendor')
                part_no = filename_parts.get('part_no', 'NoPart')
                description = filename_parts.get('description', 'NoDesc')
        
                # Clean filename parts
                vendor_code = re.sub(r'[^\w\-_]', '', str(vendor_code))[:10]
                part_no = re.sub(r'[^\w\-_]', '', str(part_no))[:15]
                description = re.sub(r'[^\w\-_]', '', str(description))[:20]
        
                filename = f"{vendor_code}_{part_no}_{description}.xlsx"
        
                # Save workbook to temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    workbook.save(tmp_file.name)
            
                    # Store row data
                    row_data = {
                        'row_index': row_idx,
                        'filename': filename,
                        'file_path': tmp_file.name,
                        'data_dict': data_dict,
                        'mapping_count': mapping_count,
                        'steps_written': steps_written,
                        'vendor_code': vendor_code,
                        'part_no': part_no,
                        'description': description,
                        'procedure_steps': filled_steps if template_procedure_steps else []
                    }
                    st.session_state.all_row_data.append(row_data)
                
                workbook.close()
                st.write(f"✅ Row {row_idx + 1} processed: {mapping_count} fields mapped, {steps_written} steps written -> {filename}")
            
            st.success(f"🎉 Successfully processed {len(data_df)} rows!")
            return True, st.session_state.all_row_data
            
        except Exception as e:
            st.error(f"❌ Error mapping template: {e}")
            st.write("📋 Traceback:", traceback.format_exc())
            return False, []
            
    def map_data_with_section_context_for_row(self, template_fields, data_df, row_idx):
        """Map data for specific row"""
        mapping_results = {}
        used_columns = set()

        try:
            data_columns = data_df.columns.tolist()

            for coord, field in template_fields.items():
                try:
                    best_match = None
                    best_score = 0.0
                    field_value = field['value']
                    section_context = field.get('section_context')

                    # Use existing mapping logic but for specific row
                    if section_context and section_context in self.section_mappings:
                        section_mappings = self.section_mappings[section_context]['field_mappings']

                        for template_field_key, data_column_pattern in section_mappings.items():
                            normalized_field_value = self.preprocess_text(field_value)
                            normalized_template_key = self.preprocess_text(template_field_key)

                            if normalized_field_value == normalized_template_key:
                                if section_context == "procedure_information":
                                    expected_column = data_column_pattern 
                                else:
                                    section_prefix = section_context.split('_')[0].capitalize()
                                    expected_column = f"{section_prefix} {data_column_pattern}".strip()

                                for data_col in data_columns:
                                    if data_col in used_columns:
                                        continue
                                    if self.preprocess_text(data_col) == self.preprocess_text(expected_column):
                                        best_match = data_col
                                        best_score = 1.0
                                        break

                                if not best_match:
                                    for data_col in data_columns:
                                        if data_col in used_columns:
                                            continue
                                        similarity = self.calculate_similarity(expected_column, data_col)
                                        if similarity > best_score and similarity >= self.similarity_threshold:
                                            best_score = similarity
                                            best_match = data_col
                                break

                    # Fallback logic (same as original)
                    if not best_match:
                        for data_col in data_columns:
                            if data_col in used_columns:
                                continue
                            similarity = self.calculate_similarity(field_value, data_col)
                            if similarity > best_score and similarity >= self.similarity_threshold:
                                best_score = similarity
                                best_match = data_col

                    mapping_results[coord] = {
                        'template_field': field_value,
                        'data_column': best_match,
                        'similarity': best_score,
                        'field_info': field,
                        'section_context': section_context,
                        'is_mappable': best_match is not None
                    }

                    if best_match:
                        used_columns.add(best_match)

                except Exception as e:
                    st.error(f"Error mapping field {coord}: {e}")
                    continue

        except Exception as e:
            st.error(f"Error in map_data_with_section_context_for_row: {e}")

        return mapping_results
    
    def write_filled_steps_to_template(self, worksheet, filled_steps):
        """Write filled procedure steps to merged cells B to P starting from Row 28"""
        try:
            from openpyxl.cell import MergedCell
            from openpyxl.styles import Font, Alignment

            print(f"\n=== WRITING FILLED PROCEDURE STEPS ===")
            st.write(f"🔄 Writing {len(filled_steps)} filled procedure steps to template")

            start_row = 28
            target_col = 2  # Column B
            end_col = 18    # Column P

            steps_written = 0

            for i, step in enumerate(filled_steps):
                step_row = start_row + i
                step_text = step.strip()
            
                # Safety check
                if step_row > worksheet.max_row + 20:
                    st.warning(f"⚠️ Stopping at row {step_row} to avoid exceeding template boundaries")
                    break
            
                try:
                    # Define the merge range for this row (B to P)
                    merge_range = f"B{step_row}:R{step_row}"
                    target_cell = worksheet.cell(row=step_row, column=target_col)
                
                    print(f"📝 Writing filled step {i + 1} to {merge_range}: {step_text[:50]}...")
                    st.write(f"📝 Step {i + 1} -> {merge_range}: {step_text[:50]}...")

                    # Unmerge any existing ranges that might conflict
                    existing_merged_ranges = []
                    for merged_range in list(worksheet.merged_cells.ranges):
                        if (merged_range.min_row <= step_row <= merged_range.max_row and
                            merged_range.min_col <= end_col and merged_range.max_col >= target_col):
                            existing_merged_ranges.append(merged_range)

                    for merged_range in existing_merged_ranges:
                        try:
                            worksheet.unmerge_cells(str(merged_range))
                            print(f"🔧 Unmerged existing range: {merged_range}")
                        except Exception as unmerge_error:
                            print(f"⚠️ Warning: Could not unmerge {merged_range}: {unmerge_error}")

                    # Clear any existing content in the range
                    for col in range(target_col, end_col + 1):
                        cell = worksheet.cell(row=step_row, column=col)
                        cell.value = None

                    # Write the step text to the first cell (B)
                    target_cell.value = step_text
                    target_cell.font = Font(name='Calibri', size=10)
                    target_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                    # Merge the cells B to P for this row
                    try:
                        worksheet.merge_cells(merge_range)
                        print(f"✅ Merged range: {merge_range}")
                    except Exception as merge_error:
                        print(f"⚠️ Warning: Could not merge {merge_range}: {merge_error}")
                        st.warning(f"Could not merge {merge_range}: {merge_error}")
  
                    # Adjust row height based on text length
                    chars_per_line = 120
                    num_lines = max(1, len(step_text) // chars_per_line + 1)
                    estimated_height = 15 + (num_lines - 1) * 15
                    worksheet.row_dimensions[step_row].height = estimated_height

                    steps_written += 1
                
                except Exception as step_error:
                    print(f"❌ Error writing step {i + 1}: {step_error}")
                    st.error(f"Error writing step {i + 1}: {step_error}")
                    continue

            print(f"\n✅ FILLED PROCEDURE STEPS COMPLETED")
            print(f"   Total steps written: {steps_written}")
        
            st.success(f"✅ Successfully wrote {steps_written} filled procedure steps to template")

            return steps_written

        except Exception as e:
            print(f"💥 Critical error in write_filled_steps_to_template: {e}")
            st.error(f"Critical error writing filled procedure steps: {e}")
            return 0

# Packaging types and procedures from reference code
PACKAGING_TYPES = [
    {
        "name": "BOX IN BOX SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Box%20in%20Box%20sensitive.png",
        "description": "Double protection for sensitive items"
    },
    {
        "name": "BOX IN BOX",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Box%20in%20box.png",
        "description": "Standard double boxing protection"
    },
    {
        "name": "CARTON BOX WITH SEPARATOR FOR ONE PART",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Cardboard%20Box%20with%20Protective%20Packing.png",
        "description": "Single item with internal separator"
    },
    {
        "name": "INDIVIDUAL NOT SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20not%20sensitive.png",
        "description": "Individual packaging for standard items"
    },
    {
        "name": "INDIVIDUAL PROTECTION FOR EACH PART MANY TYPE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20each%20part%20many%20types.png",
        "description": "Different protection for various parts"
    },
    {
        "name": "INDIVIDUAL PROTECTION FOR EACH PART",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20for%20each%20part.png",
        "description": "Uniform protection for each part"
    },
    {
        "name": "INDIVIDUAL SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20Sensitive.png",
        "description": "Individual packaging for sensitive items"
    },
    {
        "name": "MANY IN ONE TYPE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Many%20in%20one.png",
        "description": "Multiple items in single packaging"
    },
    {
        "name": "SINGLE BOX",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/88ee0796f874244af8152c681df74d352cf5359a/Single%20Box.png",
        "description": "Simple single box packaging"
    }
]

def display_packaging_grid():
    """Grid-style layout for packaging selection with consistent image sizes"""
    st.header("📦 Step 1: Select Packaging Type")
    st.markdown("Choose the most appropriate packaging type for your needs:")
    
    # Custom CSS for consistent grid image sizing
    st.markdown("""
    <style>
    .grid-image-container {
        display: flex;
        justify-content: center;
        align-items: center;
        height: 200px;
        margin-bottom: 10px;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        background-color: #fafafa;
    }
    .grid-image-container img {
        max-width: 200px;
        max-height: 200px;
        object-fit: contain;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Create a grid layout for packaging options
    cols = st.columns(3)
    
    for i, packaging in enumerate(PACKAGING_TYPES):
        with cols[i % 3]:
            # Create a container for each packaging option
            with st.container():
                # Display image with consistent sizing
                try:
                    # Use HTML container for consistent sizing
                    st.markdown(f"""
                    <div class="grid-image-container">
                        <img src="{packaging['image_url']}" alt="{packaging['name']}" />
                    </div>
                    """, unsafe_allow_html=True)
                except Exception as e:
                    # Better fallback with consistent sizing
                    st.markdown("""
                    <div class="grid-image-container">
                        <div style="text-align: center; color: #666;">
                            📦<br>Image loading...
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Title and description
                st.markdown(f"**{packaging['name']}**")
                st.write(packaging["description"])
                
                # Selection button with improved styling
                is_selected = st.session_state.get('selected_packaging_type') == packaging['name']
                button_text = "✅ Selected" if is_selected else "Select"
                
                if st.button(
                    button_text,
                    key=f"pkg_{i}", 
                    use_container_width=True,
                    type="primary" if is_selected else "secondary",
                    disabled=is_selected
                ):
                    if not is_selected:
                        st.session_state.selected_packaging_type = packaging['name']
                        st.session_state.selected_packaging_image = packaging['image_url']
                        st.success(f"Selected: {packaging['name']}")
                        navigate_to_step(2)
                        st.rerun()
            
            # Add some spacing
            st.markdown("---")

# Alternative grid method using st.image with fixed parameters
def display_packaging_grid_alternative():
    """Alternative grid-style layout using st.image with fixed dimensions"""
    st.header("📦 Step 1: Select Packaging Type")
    st.markdown("Choose the most appropriate packaging type for your needs:")
    
    # Create a grid layout for packaging options
    cols = st.columns(3)
    
    for i, packaging in enumerate(PACKAGING_TYPES):
        with cols[i % 3]:
            # Create a container for each packaging option
            with st.container():
                # Display image with fixed dimensions for consistency
                try:
                    st.image(
                        packaging["image_url"], 
                        caption=packaging["name"],
                        width=180,  # Fixed width for all images
                        use_container_width=False  # Don't use container width to maintain fixed size
                    )
                except Exception as e:
                    # Consistent fallback
                    st.info("📦 Image loading...")
                    st.write(f"**{packaging['name']}**")
                    st.caption("Image will load shortly...")
                
                # Description
                st.write(packaging["description"])
                
                # Selection button with improved styling
                is_selected = st.session_state.get('selected_packaging_type') == packaging['name']
                button_text = "✅ Selected" if is_selected else "Select"
                
                if st.button(
                    button_text,
                    key=f"pkg_alt_{i}", 
                    use_container_width=True,
                    type="primary" if is_selected else "secondary",
                    disabled=is_selected
                ):
                    if not is_selected:
                        st.session_state.selected_packaging_type = packaging['name']
                        st.session_state.selected_packaging_image = packaging['image_url']
                        st.success(f"Selected: {packaging['name']}")
                        navigate_to_step(2)
                        st.rerun()
            
            # Add some spacing
            st.markdown("---")
            
def main():
    # Header
    st.title("📦 AgiloPACK")
    st.markdown("---")
    
    # Progress indicator
    steps = [
        "Select Packaging Type",
        "Upload Template File", 
        "Upload Data File",
        "Auto-Fill Template",
        "Choose Image Option",
        "Generate Final Document"
    ]

    # Create progress bar
    progress_cols = st.columns(len(steps))
    for i, (col, step) in enumerate(zip(progress_cols, steps)):
        with col:
            if i + 1 < st.session_state.current_step:
                st.success(f"✅ {i+1}. {step}")
            elif i + 1 == st.session_state.current_step:
                st.info(f"🔄 {i+1}. {step}")
            else:
                st.write(f"⏳ {i+1}. {step}")
    
    st.markdown("---")
    
    # Step 1: Select Packaging Type
    if st.session_state.current_step == 1:
        # Display the grid layout directly
        display_packaging_grid()  # Uses HTML/CSS for consistent sizing
        
        # Show selected packaging details
        if st.session_state.get('selected_packaging_type'):
            st.markdown("### 📋 Selection Summary")
            with st.expander("✅ Selected Packaging Details", expanded=True):
                col1, col2 = st.columns([1, 2])
                with col1:
                    if st.session_state.get('selected_packaging_image'):
                        try:
                            st.image(st.session_state.selected_packaging_image, width=200)
                        except:
                            st.info("📦 Selected Package")
                with col2:
                    st.write(f"**Type:** {st.session_state.selected_packaging_type}")
                    # Find description
                    selected_pkg = next((pkg for pkg in PACKAGING_TYPES if pkg['name'] == st.session_state.selected_packaging_type), None)
                    if selected_pkg:
                        st.write(f"**Description:** {selected_pkg['description']}")
                    
                    # Action buttons
                    col2a, col2b = st.columns(2)
                    with col2a:
                        if st.button("🔄 Change Selection", type="secondary"):
                            st.session_state.selected_packaging_type = None
                            st.session_state.selected_packaging_image = None
                            st.rerun()
                    with col2b:
                        if st.button("Continue to Step 2 →", type="primary"):
                            navigate_to_step(2)
                            st.rerun()
        
    # Step 2: Upload Template File
    elif st.session_state.current_step == 2:
        st.header("📄 Step 2: Upload Template File")
        
        st.info(f"Selected Packaging Type: {st.session_state.selected_packaging_type}")
        
        uploaded_template = st.file_uploader(
            "Choose template file (Excel or Word)",
            type=['xlsx', 'xls', 'docx'],
            key="template_upload"
        )
        
        if uploaded_template is not None:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_template.name.split('.')[-1]}") as tmp_file:
                tmp_file.write(uploaded_template.getvalue())
                st.session_state.template_file = tmp_file.name
            
            st.success("✅ Template file uploaded successfully!")
            
            with st.expander("📖 Template Analysis", expanded=True):
                try:
                    mapper = EnhancedTemplateMapperWithImages()
                    template_procedure_steps = mapper.read_procedure_steps_from_template(st.session_state.template_file)
                    if template_procedure_steps:
                        st.success(f"✅ Found {len(template_procedure_steps)} procedure steps in template")
                    else:
                        st.warning("⚠️ No procedure steps found in template")
                    
                    template_fields, _ = mapper.find_template_fields_with_context_and_images(st.session_state.template_file)
                    if template_fields:
                        st.success(f"✅ Found {len(template_fields)} mappable fields in template")
                    else:
                        st.warning("⚠️ No mappable fields found in template")
                
                except Exception as e:
                    st.error(f"Error analyzing template: {e}")
            
            if st.button("Continue to Data Upload", key="continue_to_step3"):
                navigate_to_step(3)
        
        if st.button("← Go Back", key="back_from_2"):
            navigate_to_step(1)
    
    # Step 3: Upload Data File
    elif st.session_state.current_step == 3:
        st.header("📊 Step 3: Upload Data File (Excel)")
        
        uploaded_data = st.file_uploader(
            "Choose Excel data file",
            type=['xlsx', 'xls'],
            key="data_upload"
        )
        
        if uploaded_data is not None:
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_data.name.split('.')[-1]}") as tmp_file:
                tmp_file.write(uploaded_data.getvalue())
                st.session_state.data_file = tmp_file.name
            
            st.success("✅ Data file uploaded successfully!")
            
            try:
                df = pd.read_excel(st.session_state.data_file)
                st.write("Data Preview:")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"Error reading data file: {e}")
            
            if st.button("Continue to Auto-Fill", key="continue_to_step4"):
                navigate_to_step(4)
        
        if st.button("← Go Back", key="back_from_3"):
            navigate_to_step(2)
    
    # Step 4: Auto-Fill Template
    elif st.session_state.current_step == 4:
        st.header("🔄 Step 4: Auto-Fill Template")
    
        if st.session_state.mapping_completed:
            st.success(f"✅ Template auto-filling completed for {len(st.session_state.all_row_data)} rows!")
            if st.button("Continue to Image Options", key="continue_to_images"):
                navigate_to_step(5)
        else:
            if st.button("🚀 Start Auto-Fill Process", key="start_autofill", type="primary"):
                with st.spinner("🔄 Processing templates..."):
                    try:
                        mapper = EnhancedTemplateMapperWithImages()
                        success, all_row_data = mapper.map_template_with_data(
                            st.session_state.template_file,
                            st.session_state.data_file
                        )
                        if success:
                            st.session_state.mapping_completed = True
                            st.session_state.all_row_data = all_row_data
                            st.rerun()
                        else:
                            st.error("❌ Auto-fill process failed")
                    except Exception as e:
                        st.error(f"❌ Error during auto-fill: {e}")
                        st.code(traceback.format_exc())
    
        if st.button("← Go Back", key="back_from_4"):
            navigate_to_step(3)
    
    # Step 5: Choose Image Option
    elif st.session_state.current_step == 5:
        st.header("🖼️ Step 5: Choose Image Option")
        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("🔍 Smart Extract from Data File", use_container_width=True):
                st.session_state.image_option = 'extract'
                with st.spinner("🧠 Analyzing and extracting images..."):
                    extractor = EnhancedImageExtractor()
                    extracted_images = extractor.extract_images_from_excel(st.session_state.data_file)
                    if extracted_images and 'all_sheets' in extracted_images:
                        st.session_state.extracted_excel_images = extracted_images['all_sheets']
                        st.success(f"✅ Extracted {len(st.session_state.extracted_excel_images)} images!")
                    else:
                        st.warning("No images found in the Excel file.")
        with col2:
            if st.button("📁 Upload New Images", use_container_width=True):
                st.session_state.image_option = 'upload'
        with col3:
            if st.button("📄 Generate Without Images", use_container_width=True):
                st.session_state.image_option = 'no_images'

        if st.session_state.image_option == 'upload':
            st.subheader("📤 Upload Images by Type")
            image_types = ['current', 'primary', 'secondary', 'label']
            for img_type in image_types:
                uploaded_img = st.file_uploader(f"Choose {img_type} image", type=['png', 'jpg', 'jpeg'], key=f"img_upload_{img_type}")
                if uploaded_img:
                    img_bytes = uploaded_img.read()
                    img_b64 = base64.b64encode(img_bytes).decode()
                    st.session_state.uploaded_images[f"{img_type}_uploaded"] = {'data': img_b64, 'type': img_type}
                    st.image(img_bytes, width=150)

        if st.session_state.image_option:
             if st.button("Continue to Final Generation", key="continue_to_step6", type="primary"):
                navigate_to_step(6)

        if st.button("← Go Back", key="back_from_5"):
            navigate_to_step(4)

    # Step 6: Generate Final Document
    elif st.session_state.current_step == 6:
        st.header("🎨 Step 6: Generate Final Documents")
        if st.button("🚀 Generate All Templates", type="primary", use_container_width=True):
            with st.spinner("🎨 Generating templates..."):
                try:
                    extractor = EnhancedImageExtractor()
                    generated_files = []
                
                    for i, row_data in enumerate(st.session_state.all_row_data):
                        workbook = openpyxl.load_workbook(row_data['file_path'])
                        worksheet = workbook.active
                        
                        images_to_add = {}
                        if st.session_state.image_option == 'extract':
                            all_extracted = {'all_sheets': st.session_state.extracted_excel_images}
                            images_to_add = extractor.extract_images_for_part(
                                st.session_state.data_file,
                                row_data.get('part_no', ''),
                                all_extracted,
                                row_data.get('vendor_code', ''),
                                current_row=row_data['row_index'] + 2 
                            )
                        elif st.session_state.image_option == 'upload':
                            images_to_add = st.session_state.uploaded_images

                        if images_to_add:
                            _, temp_paths = extractor.add_images_to_template(worksheet, images_to_add)
                        
                        final_filename = f"Final_{row_data['filename']}"
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                            workbook.save(tmp_file.name)
                            with open(tmp_file.name, 'rb') as f:
                                file_bytes = f.read()
                            generated_files.append({'filename': final_filename, 'data': file_bytes})
                        
                        workbook.close()
                        if 'temp_paths' in locals():
                            for temp_path in temp_paths:
                                try:
                                    os.unlink(temp_path)
                                except: pass
                
                    st.success(f"🎉 Successfully generated {len(generated_files)} templates!")

                    # Download Section
                    st.subheader("📥 Download Generated Templates")
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for file_info in generated_files:
                            zip_file.writestr(file_info['filename'], file_info['data'])
                    
                    st.download_button(
                        label="📦 Download All Templates (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=f"Generated_Templates_{datetime.now().strftime('%Y%m%d')}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ Error generating templates: {e}")
                    st.code(traceback.format_exc())
        
        if st.button("← Go Back", key="back_from_6"):
            navigate_to_step(5)
    
    # Sidebar
    with st.sidebar:
        st.header("ℹ️ Help & Information")
        st.write(f"**Step**: {st.session_state.current_step}/6")
        if st.session_state.selected_packaging_type:
            st.write(f"**Packaging Type**: {st.session_state.selected_packaging_type}")
        st.subheader("Instructions")
        st.write("""
        1. **Select Type**: Choose a packaging type.
        2. **Upload Template**: Upload your Excel template.
        3. **Upload Data**: Upload your Excel data file.
        4. **Auto-Fill**: Let the AI map and fill the data.
        5. **Add Images**: Choose an image option.
        6. **Generate**: Download your final documents.
        """)
        if st.button("🔄 Reset All", type="secondary"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

if __name__ == "__main__":
    main()
